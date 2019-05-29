from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from timetable import Timetable
from spreadsheet_data.full_time_semester_event_data import *


class Room:

    def __init__(self):
        self.tools = Timetable()

    def exec_comand(self, spreadsheet_data, args):
        print("start")
        for room_key, room in spreadsheet_data.room_data.data.items():
            print(str(room.building_id) + ' - ' + str(room.name))
            if id(spreadsheet_data.full_time_first_semester_event_data) in \
                    room.referenced_by:
                for event_id in room.referenced_by[
                    id(spreadsheet_data.full_time_first_semester_event_data)]:
                    event = \
                        spreadsheet_data.full_time_first_semester_event_data \
                            .data[event_id]
                    if event.event_time != 0:
                        print(event.event_time.get_string())
                    else:
                        print('??: ??:?? - ??:??')
                    print(event.course_id)
                    print(event_types[event.event_type])
                    print(event.trainer_id)
                    print()
            print("----------------------------")
        # for item in exist_room_date:
        #     print(item[self.tools.columnNameToIndex['sala']].value)

# workbook = load_workbook(filename='data/data.xlsx')
# sheet = workbook['zima-s']
#
# timetable = Timetable()
# # rooms = tools.get_with_not_null_column(tools.columnNameToIndex['sala'], sheet)
# # print(rooms[0][tools.columnNameToIndex['sala']].value)
# all_rooms = timetable.get_unique_value_from_column('sala', sheet)
# print(all_rooms)

#
# i=0
# for row in sheet.rows:
#     # print(row[columnNameToIndex['studia']].value)
#     for cell in row:
#         print(cell.value)
#     break

# first_column = sheet_ranges['D']
# print(*sheet_ranges)
# # # Print the contents
# for x in range(len(first_column)):
#     print(first_column[x].value)
